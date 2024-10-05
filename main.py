from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from PIL import Image, ImageTk
import time
import openpyxl
import csv
import matplotlib.pyplot as plt
import mplcursors
import io

class popup:
    def address_data_popup(raw_address_hex):
        address_data = Toplevel(root)
        address_data.title("Address Data Operation Report")
        address_data.geometry("1280x400")
        address_data.resizable(0,0)
        
        address_data_frame = Text(address_data, wrap=WORD, width=100, height=4, font="Arial 14 bold", padx=10, pady=10)
        address_data_frame.pack(padx=5, pady=5)
        
        text_to_insert = (f"{month} {date},{year}  {hour}:{minute}:{second} \n"
                        f"Operator Name: {operator_name.get()} \n"
                        f"Room Temperature: {room_temperature.get()} \nRoom Humidity: {room_humidity.get()} ")
        address_data_frame.insert(INSERT, text_to_insert)
        address_data_frame.config(state="disabled")
        
        raw_address_data_frame = LabelFrame(address_data, text="Raw Address Data", fg ="#00D2FF",padx=5, pady=5)
        raw_address_data_frame.pack()
        
        Label(raw_address_data_frame,text=raw_address_hex, wraplength=800, font="Arial 14 bold", width=100).pack()
        
        Label(address_data,text="\n Address Data Parsing", font="Arial 18 bold", fg ="#00D2FF").pack(anchor="nw", padx=5, pady=5)
        
        temp_address_hex = raw_address_hex
        
        raw_address_hex = raw_address_hex.replace(" ","")
        Data_date= raw_address_hex[0:10]
        Data_time = raw_address_hex[10:18]
        HK_DATA_ADDRESS = raw_address_hex[40:48]
        SAT_LOG_ADDRESS = raw_address_hex[48:56]
        IMG_INFO_ADDRESS = raw_address_hex[56:64]
        MSN_2_ADDRESS = raw_address_hex[64:72]
        MSN_3_ADDRESS = raw_address_hex[72:80]
        MSN_1_ADDRESS = raw_address_hex[80:88]
        RSV_TABLE_ADDRESS = raw_address_hex[88:96]
        Flag_ADDRESS= raw_address_hex[96:104]
        ADDRESS_COUNTER = int(raw_address_hex[104:112],16)
        
        headers = ["Date", "Time", "HK_DATA_ADDR", "SAT_LOG_ADDR", "IMG_INFO_ADDR",
           "MSN_2_ADDR", "MSN_3_ADDR", "MSN_1_ADDR", "Flag_ADDR", "RSV_TABLE_ADDR", "ADDR_COUNTER"]
        
        data = [Data_date, Data_time, HK_DATA_ADDRESS, SAT_LOG_ADDRESS, IMG_INFO_ADDRESS, MSN_2_ADDRESS, 
                MSN_3_ADDRESS, MSN_1_ADDRESS, RSV_TABLE_ADDRESS, Flag_ADDRESS, ADDRESS_COUNTER]
        
        tree = ttk.Treeview(address_data, columns=headers, show="headings", height=2, padding=5)
        
        # Add headers to the treeview
        for header in headers:
            tree.heading(header, text=header)
            tree.column(header, anchor=CENTER, stretch=TRUE)
        
        # Insert data into the treeview
        tree.insert("", END, values=data)
        tree.pack(expand=TRUE)

        save_btn = Button(address_data, text = 'Save Address Data', command=lambda:address_save(text_to_insert, temp_address_hex, headers, data, address_data))
        save_btn.pack(padx=10, pady=10, side=TOP)
        
    def beacon_data_popup(raw_beacon_hex):
        beacon_data = Toplevel(root)
        beacon_data.title("Beacon Data Operation Report")
        beacon_data.geometry("1280x600")
        beacon_data.resizable(0,0)
        
        beacon_data_frame = Text(beacon_data, wrap=WORD, width=100, height=4, font="Arial 14 bold", padx=10, pady=10)
        beacon_data_frame.pack(padx=5, pady=5)
        
        text_to_insert = (f"{month} {date},{year}  {hour}:{minute}:{second} \n"
                        f"Operator Name: {operator_name.get()} \n"
                        f"Room Temperature: {room_temperature.get()} \nRoom Humidity: {room_humidity.get()} ")
        beacon_data_frame.insert(INSERT, text_to_insert)
        beacon_data_frame.config(state="disabled")
        
        raw_beacon_data_frame = LabelFrame(beacon_data, text="Raw Address Data", fg ="#00D2FF",padx=5, pady=5)
        raw_beacon_data_frame.pack()
        
        Label(raw_beacon_data_frame,text=raw_beacon_hex, wraplength=800, font="Arial 14 bold", width=100).pack()
        
        Label(beacon_data,text="\n Beacon Data Parsing", font="Arial 18 bold", fg ="#00D2FF").pack(anchor="nw", padx=5, pady=5)
        
        temp_beacon_hex = raw_beacon_hex

        raw_beacon_hex = raw_beacon_hex.replace(" ","")
        Data_date= raw_beacon_hex[0:10]
        Data_time = raw_beacon_hex[10:18]
        Header = raw_beacon_hex[40:42]
        Beacon_Type = raw_beacon_hex[43:44]
        if(Beacon_Type == "1"):
            SAT_Day = int(str(raw_beacon_hex[44:46]) + str(raw_beacon_hex[42:43]),16)
            SAT_HRS = int(raw_beacon_hex[46:48],16)
            BAT_V = int(str(raw_beacon_hex[50:52]) + str(raw_beacon_hex[48:50]),16)/1000
            BAT_C = int(str(raw_beacon_hex[54:56]) + str(raw_beacon_hex[52:54]),16)/1000
            BAT_T =int(str(raw_beacon_hex[58:60]) + str(raw_beacon_hex[56:58]),16)/100
            RAW_C = int(str(raw_beacon_hex[60:62]),16)/10
            SOL_TOT_V = int(str(raw_beacon_hex[64:66]) + str(raw_beacon_hex[62:64]),16)/1000
            SOL_TOT_C =int(str(raw_beacon_hex[68:70]) + str(raw_beacon_hex[66:68]),16)/1000
            ANT_P_T = int(str(raw_beacon_hex[70:72]),16)
            BPB_T = int(str(raw_beacon_hex[72:74]),16)
            OBC_T = int(str(raw_beacon_hex[74:76]),16)
            SOL_P1_T = int(str(raw_beacon_hex[76:78]),16)
            SOL_P2_T = int(str(raw_beacon_hex[78:80]),16)
            SOL_P3_T = int(str(raw_beacon_hex[80:82]),16)
            SOL_P4_T = int(str(raw_beacon_hex[82:84]),16)
            SOL_P5_T = int(str(raw_beacon_hex[84:86]),16)
            SOL_MSN_STAT = str(bin(int(str(raw_beacon_hex[86:88]),16)))
            MSN3_STAT= SOL_MSN_STAT[2:3]
            MSN2_STAT= SOL_MSN_STAT[3:4]
            MSN1_STAT= SOL_MSN_STAT[4:5]
            SOL_P5_STAT= SOL_MSN_STAT[5:6]
            SOL_P4_STAT= SOL_MSN_STAT[6:7]
            SOL_P3_STAT= SOL_MSN_STAT[7:8]
            SOL_P2_STAT= SOL_MSN_STAT[8:9]
            SOL_P1_STAT= SOL_MSN_STAT[9:10]

            ANT_STAT = int(str(raw_beacon_hex[88:89]),16)
            UP_STAT = int(str(raw_beacon_hex[89:90]),16)
            OPERATION_MODE = int(str(raw_beacon_hex[90:92]),16)
            if OPERATION_MODE == 90:
                OPERATION_MODE="NORMAL"
            elif OPERATION_MODE == 106:
                OPERATION_MODE="LOW_POWER"
            elif OPERATION_MODE == 122:
                OPERATION_MODE == "SAFE_MODE"
            OBC_RST_COUNT = int(str(raw_beacon_hex[94:96]) + str(raw_beacon_hex[92:94]),16)
            RST_RST_COUNT =int(str(raw_beacon_hex[98:100])+str(raw_beacon_hex[96:98]),16)
            LAST_RST =int(str(raw_beacon_hex[100:102]),16)
            CHK_CRC = str(raw_beacon_hex[102:104])
            
            headers =["Date", "Time", "Header", "Beacon_Type", "SAT_Day", "SAT_HRS", "BAT_V", "BAT_C", "BAT_T",
                      "RAW_C", "SOL_TOT_V", "SOL_TOT_C", "ANT_P_T", "BPB_T", "OBC_T", "SOL_P1_T", "SOL_P2_T", "SOL_P3_T",
                      "SOL_P4_T", "SOL_P5_T", "MSN3_STAT", "MSN2_STAT", "MSN1_STAT", "SOL_P5_STAT", "SOL_P4_STAT", "SOL_P3_STAT", "SOL_P2_STAT",
                      "SOL_P1_STAT","ANT_STAT", "UP_STAT", "OPERATION_MODE", "OBC_RST_COUNT", "RST_RST_COUNT", "LAST_RST", "CHK_CRC"]
           
            headers_1 = ["Date", "Time", "Header", "Beacon_Type", "SAT_Day", "SAT_HRS", "BAT_V", "BAT_C", "BAT_T"]
            headers_2 = ["RAW_C", "SOL_TOT_V", "SOL_TOT_C", "ANT_P_T", "BPB_T", "OBC_T", "SOL_P1_T", "SOL_P2_T", "SOL_P3_T"]
            headers_3 = ["SOL_P4_T", "SOL_P5_T", "MSN3_STAT", "MSN2_STAT", "MSN1_STAT", "SOL_P5_STAT", "SOL_P4_STAT", "SOL_P3_STAT", "SOL_P2_STAT"]
            headers_4 = ["SOL_P1_STAT","ANT_STAT", "UP_STAT", "OPERATION_MODE", "OBC_RST_COUNT", "RST_RST_COUNT", "LAST_RST", "CHK_CRC"]
        
            data = [Data_date, Data_time, Header, Beacon_Type, SAT_Day, SAT_HRS, BAT_V, BAT_C, BAT_T,
                    RAW_C, SOL_TOT_V, SOL_TOT_C, ANT_P_T, BPB_T, OBC_T, SOL_P1_T, SOL_P2_T, SOL_P3_T,
                    SOL_P4_T, SOL_P5_T, MSN3_STAT, MSN2_STAT, MSN1_STAT, SOL_P5_STAT, SOL_P4_STAT, SOL_P3_STAT, SOL_P2_STAT,
                    SOL_P1_STAT, ANT_STAT, UP_STAT, OPERATION_MODE, OBC_RST_COUNT, RST_RST_COUNT, LAST_RST, CHK_CRC]
            
            data_1 = [Data_date, Data_time, Header, Beacon_Type, SAT_Day, SAT_HRS, BAT_V, BAT_C, BAT_T]
            data_2 = [RAW_C, SOL_TOT_V, SOL_TOT_C, ANT_P_T, BPB_T, OBC_T, SOL_P1_T, SOL_P2_T, SOL_P3_T]
            data_3 = [SOL_P4_T, SOL_P5_T, MSN3_STAT, MSN2_STAT, MSN1_STAT, SOL_P5_STAT, SOL_P4_STAT, SOL_P3_STAT, SOL_P2_STAT]
            data_4 = [SOL_P1_STAT, ANT_STAT, UP_STAT, OPERATION_MODE, OBC_RST_COUNT, RST_RST_COUNT, LAST_RST, CHK_CRC]

            tree = ttk.Treeview(beacon_data, columns=headers_1, show="headings", height=2)
            tree_2 = ttk.Treeview(beacon_data, columns=headers_2, show="headings", height=2)
            tree_3 = ttk.Treeview(beacon_data, columns=headers_3, show="headings", height=2)
            tree_4 = ttk.Treeview(beacon_data, columns=headers_4, show="headings", height=2)
            # Add headers to the treeview
            for header in headers_1:
                tree.heading(header, text=header)
                tree.column(header, anchor=CENTER)
            
            for header in headers_2:
                tree_2.heading(header, text=header)
                tree_2.column(header, anchor=CENTER)
            
            for header in headers_3:
                tree_3.heading(header, text=header)
                tree_3.column(header, anchor=CENTER)
            
            for header in headers_4:
                tree_4.heading(header, text=header)
                tree_4.column(header, anchor=CENTER)
            
            # Insert data into the treeview
            tree.insert("", END, values=data_1)
            tree_2.insert("", END, values = data_2)
            tree_3.insert("", END, values=data_3)
            tree_4.insert("", END, values = data_4)
            
            tree.pack(fill="both", expand=TRUE)
            tree_2.pack(fill="both", expand=TRUE)
            tree_3.pack(fill="both", expand=TRUE)
            tree_4.pack(fill="both", expand=TRUE)
            
            save_btn = Button(beacon_data, text = 'Save Beacon Type_1', command=lambda:beacon_save(text_to_insert, temp_beacon_hex, headers, data, beacon_data, Beacon_Type))
            save_btn.pack(padx=5, pady=5, side=TOP)
            
        elif(Beacon_Type == "2"):
            SAT_Day =   int(str(raw_beacon_hex[44:46]) + str(raw_beacon_hex[42:43]),16)    
            SOL_P1_V =  int(str(raw_beacon_hex[46:48]),16)/10
            SOL_P2_V =  int(str(raw_beacon_hex[48:50]),16)/10
            SOL_P3_V =  int(str(raw_beacon_hex[50:52]),16)/10
            SOL_P4_V =  int(str(raw_beacon_hex[52:54]),16)/10
            SOL_P5_V =  int(str(raw_beacon_hex[54:56]),16)/10
            SOL_P1_C =  int(str(raw_beacon_hex[56:58]),16)/10
            SOL_P2_C = int(str(raw_beacon_hex[58:60]),16)/10
            SOL_P3_C = int(str(raw_beacon_hex[60:62]),16)/10
            SOL_P4_C = int(str(raw_beacon_hex[62:64]),16)/10
            SOL_P5_C = int(str(raw_beacon_hex[64:66]),16)/10
            GYRO_X = int(str(raw_beacon_hex[68:70]) + str(raw_beacon_hex[66:68]),16) 
            GYRO_Y = int(str(raw_beacon_hex[72:74]) + str(raw_beacon_hex[70:72]),16)
            GYRO_Z = int(str(raw_beacon_hex[76:78]) + str(raw_beacon_hex[74:76]),16)
            ACCL_X = int(str(raw_beacon_hex[80:82]) + str(raw_beacon_hex[78:80]),16)
            ACCL_Y = int(str(raw_beacon_hex[84:86]) + str(raw_beacon_hex[82:84]),16)
            ACCL_Z = int(str(raw_beacon_hex[88:90]) + str(raw_beacon_hex[86:88]),16)
            MAG_X = int(str(raw_beacon_hex[92:94]) + str(raw_beacon_hex[90:92]),16) 
            MAG_Y = int(str(raw_beacon_hex[96:98]) + str(raw_beacon_hex[94:96]),16) 
            MAG_Z = int(str(raw_beacon_hex[100:102]) + str(raw_beacon_hex[98:100]),16)
            CHK_CRC = str(raw_beacon_hex[102:104])
            
            headers = ["Date", "Time", "Header", "Beacon_Type", "SAT_Day", "SOL_P1_V", "SOL_P2_V", "SOL_P3_V",
                       "SOL_P4_V", "SOL_P5_V", "SOL_P1_C", "SOL_P2_C", "SOL_P3_C", "SOL_P4_C", "SOL_P5_C", "GYRO_X",
                       "GYRO_Y", "GYRO_Z", "ACCL_X", "ACCL_Y", "ACCL_Z", "MAG_X", "MAG_Y", "MAG_Z", "CHK_CRC"]
            
            headers_1 = ["Date", "Time", "Header", "Beacon_Type", "SAT_Day", "SOL_P1_V", "SOL_P2_V", "SOL_P3_V"]
            headers_2 = ["SOL_P4_V", "SOL_P5_V", "SOL_P1_C", "SOL_P2_C", "SOL_P3_C", "SOL_P4_C", "SOL_P5_C", "GYRO_X"]
            headers_3 = ["GYRO_Y", "GYRO_Z", "ACCL_X", "ACCL_Y", "ACCL_Z", "MAG_X", "MAG_Y", "MAG_Z", "CHK_CRC"]
            
            data = [Data_date, Data_time, Header, Beacon_Type, SAT_Day, SOL_P1_V, SOL_P2_V, SOL_P3_V,
                    SOL_P4_V, SOL_P5_V, SOL_P1_C, SOL_P2_C, SOL_P3_C, SOL_P4_C, SOL_P5_C, GYRO_X,
                    GYRO_Y, GYRO_Z, ACCL_X, ACCL_Y, ACCL_Z, MAG_X, MAG_Y, MAG_Z,CHK_CRC]
            
            data_1 = [Data_date, Data_time, Header, Beacon_Type, SAT_Day, SOL_P1_V, SOL_P2_V, SOL_P3_V]
            data_2 = [SOL_P4_V, SOL_P5_V, SOL_P1_C, SOL_P2_C, SOL_P3_C, SOL_P4_C, SOL_P5_C, GYRO_X]
            data_3 = [GYRO_Y, GYRO_Z, ACCL_X, ACCL_Y, ACCL_Z, MAG_X, MAG_Y, MAG_Z,CHK_CRC]
            
            tree = ttk.Treeview(beacon_data, columns=headers_1, show="headings", height=2)
            tree_2 = ttk.Treeview(beacon_data, columns=headers_2, show="headings", height=2)
            tree_3 = ttk.Treeview(beacon_data, columns=headers_3, show="headings", height=2)

            # Add headers to the treeview
            for header in headers_1:
                tree.heading(header, text=header)
                tree.column(header, anchor=CENTER)
            
            for header in headers_2:
                tree_2.heading(header, text=header)
                tree_2.column(header, anchor=CENTER)
            
            for header in headers_3:
                tree_3.heading(header, text=header)
                tree_3.column(header, anchor=CENTER)

            # Insert data into the treeview
            tree.insert("", END, values=data_1)
            tree_2.insert("", END, values = data_2)
            tree_3.insert("", END, values=data_3)
            
            tree.pack(fill="both", expand=TRUE)
            tree_2.pack(fill="both", expand=TRUE)
            tree_3.pack(fill="both", expand=TRUE)
            
            save_btn = Button(beacon_data, text = 'Save Beacon Type_2', command=lambda:beacon_save(text_to_insert, temp_beacon_hex, headers, data, beacon_data, Beacon_Type))
            save_btn.pack(padx=5, pady=5, side=TOP) 

    def hk_data_popup(raw_hk_hex):
        hk_data = Toplevel(root)
        hk_data.title("House Keeping Data Operation Report")
        hk_data.geometry("1280x730")
        hk_data.resizable(0,0)
        
        hk_hex_frame = Text(hk_data, wrap=WORD, width=100, height=4, font="Arial 14 bold", padx=10, pady=10)
        hk_hex_frame.pack(padx=5, pady=5)
        
        text_to_insert = (f"{month} {date},{year}  {hour}:{minute}:{second} \n"
                        f"Operator Name: {operator_name.get()} \n"
                        f"Room Temperature: {room_temperature.get()} \nRoom Humidity: {room_humidity.get()} ")
        hk_hex_frame.insert(INSERT, text_to_insert)
        hk_hex_frame.config(state="disabled")
        
        raw_hk_hex_frame = LabelFrame(hk_data, text="Raw Housekeeping Data", fg ="#00D2FF",padx=5, pady=5)
        raw_hk_hex_frame.pack()
        
        canvas = Canvas(raw_hk_hex_frame, width=820, height=60)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        
        # Add a scrollbar to the canvas
        scrollbar = Scrollbar(raw_hk_hex_frame, command=canvas.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        canvas.config(yscrollcommand=scrollbar.set)
        
        # Create a frame inside the canvas
        frame = Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor=NW)

        Label(frame,text=raw_hk_hex, wraplength=800, font="Arial 14 bold", width=100).pack()
        
        # Update the scroll region
        frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))
        
        Label(hk_data,text="\n HK Data Parsing", font="Arial 18 bold", fg ="#00D2FF").pack(anchor="nw", padx=5, pady=5)
        
        hk_data_array = raw_hk_hex.strip().split("\n")
        
        print(f"raw array data--> {hk_data_array}")
        all_data_1 = ""
        for line in hk_data_array:
            data = line.replace(" ","")
            data_1 = data[40:-4]
            print(f"individual Data: --> {data_1}\n")
            all_data_1 += data_1  # Concatenating each data_1 value to the variable all_data_1

        print(f"Actual Data: --> {all_data_1}")
                          
        data_with_newline = all_data_1.replace("feba", "fe\nba")
        array_with_newline = data_with_newline.split("\n")
        print(f"Actual Array Data -> \n{array_with_newline}")
        
        # Counting characters in each element of the array
        character_counts = [len(element) for element in array_with_newline]
        for index, count in enumerate(character_counts, start=1):
            print(f"Character count in index {index}: {count}")

        headers = ["Day", "Hour", "Min","Sec",
                      "SOLAR_P1_CS","SOLAR_P2_CS","SOLAR_P3_CS","SOLAR_P4_CS","SOLAR_P5_CS","SOLAR_TOTAL_CS",
                      "BATT_CS","RAW_CS","RESET_3V3_CS","MSN_3V3_1_CS","MSN_3V3_2_CS",
                      "MSN_5V_CS","UNREG1_CS","UNREG2_CS","SP1_VOLT","SP2_VOLT","SP3_VOLT","SP4_VOLT",
                      "SP5_VOLT","SOLAR_TOTAL_VOLT","RAW_VOLTS","BATT_VOLTS","Y_TEMP_N","Y_TEMP_P",
                      "Z_TEMP_N","Z_TEMP_P","X_TEMP_N","X_TEMP_P","BPB_TEMP","BATT_TEMP","OBC_INT_TEMP",
                      "GYRO_X","GYRO_Y","GYRO_Z","MAG_X","MAG_Y","MAG_Z"]
        
        headers_1 = ["Day", "Hour", "Min","Sec","SOLAR_P1_CS","SOLAR_P2_CS","SOLAR_P3_CS"]
        headers_2 = ["SOLAR_P4_CS","SOLAR_P5_CS","SOLAR_TOTAL_CS","BATT_CS","RAW_CS","RESET_3V3_CS","MSN_3V3_1_CS"]
        headers_3 = ["MSN_3V3_2_CS","MSN_5V_CS","UNREG1_CS","UNREG2_CS","SP1_VOLT","SP2_VOLT","SP3_VOLT"]
        headers_4 = ["SP4_VOLT","SP5_VOLT","SOLAR_TOTAL_VOLT","RAW_VOLTS","BATT_VOLTS","Y_TEMP_N","Y_TEMP_P"]
        headers_5 = ["Z_TEMP_N","Z_TEMP_P","X_TEMP_N","X_TEMP_P","BPB_TEMP","BATT_TEMP","OBC_INT_TEMP"]
        headers_6 = ["GYRO_X","GYRO_Y","GYRO_Z","MAG_X","MAG_Y","MAG_Z"]
        
        data_array = []
        
        if all_data_1:
            button_frame = Frame(hk_data)
            button_frame.pack(side=BOTTOM, pady=10)

            save_btn = Button(button_frame, text = 'Save Housekeeping Data', command=lambda:hk_save(text_to_insert, array_with_newline, headers, data_array, hk_data))
            save_btn.pack(padx=5, pady=5, side=LEFT)

            show_btn = Button(button_frame, text = 'Show HK Chart', command=lambda:hk_chart(headers, data_array))
            show_btn.pack(padx=5, pady=5, side=LEFT)

        tree = ttk.Treeview(hk_data, columns=headers_1, show="headings", height=2)
        tree_2 = ttk.Treeview(hk_data, columns=headers_2, show="headings", height=2)
        tree_3 = ttk.Treeview(hk_data, columns=headers_3, show="headings", height=2)
        tree_4 = ttk.Treeview(hk_data, columns=headers_4, show="headings", height=2)
        tree_5 = ttk.Treeview(hk_data, columns=headers_5, show="headings", height=2)
        tree_6 = ttk.Treeview(hk_data, columns=headers_6, show="headings", height=2)

        # Add headers to the treeview
        for header in headers_1:
            tree.heading(header, text=header)
            tree.column(header, anchor=CENTER)
            
        for header in headers_2:
            tree_2.heading(header, text=header)
            tree_2.column(header, anchor=CENTER)
            
        for header in headers_3:
            tree_3.heading(header, text=header)
            tree_3.column(header, anchor=CENTER)
            
        for header in headers_4:
            tree_4.heading(header, text=header)
            tree_4.column(header, anchor=CENTER)
            
        for header in headers_5:
            tree_5.heading(header, text=header)
            tree_5.column(header, anchor=CENTER)
            
        for header in headers_6:
            tree_6.heading(header, text=header)
            tree_6.column(header, anchor=CENTER)
            
        for result in array_with_newline:
            Day = int(result[2:4],16)
            Hour = int(result[4:6],16)
            Min = int(result[6:8],16)
            Sec = int(result[8:10],16)  
            SOLAR_P1_CS = int(result[10:14],16)
            SOLAR_P2_CS = int(result[14:18],16)
            SOLAR_P3_CS = int(result[18:22],16)       
            SOLAR_P4_CS = int(result[22:26],16)
            SOLAR_P5_CS = int(result[26:30],16)
            SOLAR_TOTAL_CS =str(result[30:34])
            SOLAR_TOTAL_CS = twosComplement_hex(SOLAR_TOTAL_CS)
            BATT_CS = str(result[34:38])  
            BATT_CS = twosComplement_hex(BATT_CS)     
            RAW_CS = str(result[38:42])     
            RAW_CS = twosComplement_hex(RAW_CS)          
            RESET_3V3_CS = int(result[42:46],16)
            MSN_3V3_1_CS = int(result[46:50],16)            
            MSN_3V3_2_CS = int(result[50:54],16)
            MSN_5V_CS = int(result[54:58],16)            
            UNREG1_CS = int(result[58:62],16)
            UNREG2_CS = int(result[62:66],16)            
            SP1_VOLT = int(result[66:70],16)
            SP2_VOLT = int(result[70:74],16)                        
            SP3_VOLT = int(result[74:78],16)
            SP4_VOLT = int(result[78:82],16)
            SP5_VOLT = int(result[82:86],16)
            SOLAR_TOTAL_VOLT = int(result[86:90],16)       
            RAW_VOLTS = int(result[90:94],16)            
            BATT_VOLTS = int(result[94:98],16)            
            Y_TEMP_N = int(result[98:102],16)          
            Y_TEMP_P = int(result[102:106],16)
            Z_TEMP_N = int(result[106:110],16)
            Z_TEMP_P = int(result[110:114],16)
            X_TEMP_N = int(result[114:118],16)
            X_TEMP_P = int(result[118:122],16)
            BPB_TEMP = int(result[122:126],16)
            BATT_TEMP = int(result[126:130],16)
            OBC_INT_TEMP = int(result[130:134],16)
            GYRO_X = int(result[134:138],16)/100
            GYRO_Y= int(result[138:142],16)/100
            GYRO_Z = int(result[142:146],16)/100
            MAG_X = int(result[146:150],16)
            MAG_Y = int(result[150:154],16)
            MAG_Z = int(result[154:158],16)
            
            data= [Day, Hour, Min,Sec,
                   SOLAR_P1_CS,SOLAR_P2_CS,SOLAR_P3_CS,SOLAR_P4_CS,SOLAR_P5_CS,SOLAR_TOTAL_CS,
                   BATT_CS,RAW_CS,RESET_3V3_CS,MSN_3V3_1_CS,MSN_3V3_2_CS,
                   MSN_5V_CS,UNREG1_CS,UNREG2_CS,SP1_VOLT,SP2_VOLT,SP3_VOLT,SP4_VOLT,
                   SP5_VOLT,SOLAR_TOTAL_VOLT,RAW_VOLTS,BATT_VOLTS,Y_TEMP_N,Y_TEMP_P,
                   Z_TEMP_N,Z_TEMP_P,X_TEMP_N,X_TEMP_P,BPB_TEMP,BATT_TEMP,OBC_INT_TEMP,
                   GYRO_X,GYRO_Y,GYRO_Z,MAG_X,MAG_Y,MAG_Z]
            
            data_1 = [Day, Hour, Min,Sec,SOLAR_P1_CS,SOLAR_P2_CS,SOLAR_P3_CS]
            data_2 = [SOLAR_P4_CS,SOLAR_P5_CS,SOLAR_TOTAL_CS,BATT_CS,RAW_CS,RESET_3V3_CS,MSN_3V3_1_CS]
            data_3 = [MSN_3V3_2_CS,MSN_5V_CS,UNREG1_CS,UNREG2_CS,SP1_VOLT,SP2_VOLT,SP3_VOLT]
            data_4 = [SP4_VOLT,SP5_VOLT,SOLAR_TOTAL_VOLT,RAW_VOLTS,BATT_VOLTS,Y_TEMP_N,Y_TEMP_P]
            data_5 = [Z_TEMP_N,Z_TEMP_P,X_TEMP_N,X_TEMP_P,BPB_TEMP,BATT_TEMP,OBC_INT_TEMP]
            data_6 = [GYRO_X,GYRO_Y,GYRO_Z,MAG_X,MAG_Y,MAG_Z]

            print(f"Data -> \n{data}")
            data_array.append(data)
            
             # Insert data into the treeview
            tree.insert("", END, values=data_1)
            tree_2.insert("", END, values = data_2)
            tree_3.insert("", END, values=data_3)
            tree_4.insert("", END, values = data_4)
            tree_5.insert("", END, values=data_5)
            tree_6.insert("", END, values = data_6)
            
            tree.pack(fill="both", expand=TRUE)
            tree_2.pack(fill="both", expand=TRUE)
            tree_3.pack(fill="both", expand=TRUE)
            tree_4.pack(fill="both", expand=TRUE)
            tree_5.pack(fill="both", expand=TRUE)
            tree_6.pack(fill="both", expand=TRUE)
    
    def image_data_popup(raw_image_hex):
        image_data = Toplevel(root)
        image_data.title("Image Data Operation Report")
        image_data.geometry("1280x530")
        image_data.resizable(0,0)
        
        image_hex_frame = Text(image_data, wrap=WORD, width=100, height=4, font="Arial 14 bold", padx=10, pady=10)
        image_hex_frame.pack(padx=5, pady=5)
        
        text_to_insert = (f"{month} {date},{year}  {hour}:{minute}:{second} \n"
                      f"Operator Name: {operator_name.get()} \n"
                      f"Room Temperature: {room_temperature.get()} \nRoom Humidity: {room_humidity.get()} ")
        image_hex_frame.insert(INSERT, text_to_insert)
        image_hex_frame.config(state="disabled")
        
        raw_image_hex_frame = LabelFrame(image_data, text="Raw Image Data", fg ="#00D2FF",padx=5, pady=5)
        raw_image_hex_frame.pack()
        
        canvas = Canvas(raw_image_hex_frame, width=820, height=60)
        canvas.pack(side=LEFT, fill=BOTH, expand=True)
        
        # Add a scrollbar to the canvas
        scrollbar = Scrollbar(raw_image_hex_frame, command=canvas.yview)
        scrollbar.pack(side=RIGHT, fill=Y)
        canvas.config(yscrollcommand=scrollbar.set)
        
        # Create a frame inside the canvas
        frame = Frame(canvas)
        canvas.create_window((0, 0), window=frame, anchor=NW)

        Label(frame,text=raw_image_hex, wraplength=800, font="Arial 14 bold", width=100).pack()
    
        # Update the scroll region
        frame.update_idletasks()
        canvas.config(scrollregion=canvas.bbox("all"))
        
        Label(image_data,text="\n Image Data Parsing", font="Arial 18 bold", fg ="#00D2FF").pack(anchor="nw", padx=5, pady=5)
        
        image_data_array = raw_image_hex.strip().split("\n")
        
        print(f"raw array data--> {image_data_array}")
        all_data_1 = ""
        for line in image_data_array:
            data = line.replace(" ","")
            data_1 = data[40:-5]
            print(f"individual Data: --> {data_1}\n")
            all_data_1 += data_1  # Concatenating each data_1 value to the variable all_data_1

        print(f"Actual Data: --> {all_data_1}")
        
        data_with_newline = all_data_1.replace("ffd9", "ffd9\n")
        array_with_newline = data_with_newline.split("\n")
        print(f"Actual Image Data -> \n{array_with_newline[0]}")
        
        actual_image_hex_frame = LabelFrame(image_data, text="Actual Image Data",padx=5, pady=5)
        actual_image_hex_frame.pack()
        
        canvas_1 = Canvas(actual_image_hex_frame, width=900, height=150)
        canvas_1.pack(side=LEFT, fill=BOTH, expand=True)
        
        # Add a scrollbar to the canvas
        scrollbar_1 = Scrollbar(actual_image_hex_frame, command=canvas_1.yview)
        scrollbar_1.pack(side=RIGHT, fill=Y)
        canvas_1.config(yscrollcommand=scrollbar_1.set)
        
        # Create a frame inside the canvas
        frame_1 = Frame(canvas_1)
        canvas_1.create_window((0, 0), window=frame_1, anchor=NW)

        Label(frame_1,text=array_with_newline[0], wraplength=900, font="Arial 14 bold").pack()
    
        # Update the scroll region
        frame_1.update_idletasks()
        canvas_1.config(scrollregion=canvas_1.bbox("all"))
        
        if array_with_newline[0]:
            button_frame = Frame(image_data)
            button_frame.pack(pady=10)

            array_with_newline[0] = raw_image_hex
            
            save_btn = Button(button_frame, text = 'Save Image Data', command=lambda:image_data_save(text_to_insert, array_with_newline[0]))
            save_btn.pack(padx=5, pady=5, side=LEFT)
        
######## twos complement function ########
def twosComplement_hex(hexval):
    bits = 16
    val = int(hexval, bits)
    if val & (1 << (bits-1)):
        val -= 1 << bits
    return val

def address_save(profile, raw_address_hex, header, data, address_data):
    print(f"{profile}\n{raw_address_hex}\n{header}\n{data}")
    
    # Create or load the Excel file
    file_name = "saved_data/DATA_ADDRESS_PKT_PARSER.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if header:
            header_row = ["SN","Remarks","Raw Address Data"] + header
            sheet.append(header_row)
    # Calculate next row and insert values
    next_row = sheet.max_row + 1

    # Write formula in column A and data in column B
    if next_row == 2:  # For the second row, insert the initial number '1'
        sheet.cell(row=next_row, column=1).value = 1
    else:  # For subsequent rows, use the formula
        sheet.cell(row=next_row, column=1).value = f"=A{next_row-1}+1"

    sheet.cell(row=next_row, column=2).value = profile 
    sheet.cell(row=next_row, column=3).value = raw_address_hex
    
    # Insert data into the same row as raw_address_hex
    for i, value in enumerate(data, start=4):  # Starting from column D (4th column)
        sheet.cell(row=next_row, column=i).value = value

    # Save the Excel file
    workbook.save(file_name)
    
    # Opening a output csv file in write mode
    OutputCsvFile = csv.writer(open("saved_data/DATA_ADDRESS_PKT_PARSER_CSV.csv", 'w'), delimiter=",")

    # Traversing in each row of the worshsheet
    for eachrow in sheet.rows:
        # Writing data of the excel file into the result csv file row-by-row
        OutputCsvFile.writerow([cell.value for cell in eachrow])
    
    address_data.destroy()
    show_alert()
    
def beacon_save(profile, raw_beacon_hex, header, data, beacon_data, Beacon_Type):
    print(f"{profile}\n{raw_beacon_hex}\n{header}\n{data}")
    
    if(Beacon_Type == "1"):
        # Create or load the Excel file
        file_name = "saved_data/Beacon_Type_1.xlsx"
    elif(Beacon_Type == "2"):
        file_name = "saved_data/Beacon_Type_2.xlsx"
    
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if header:
            if(Beacon_Type == "1"):
                header_row = ["SN","Remarks","Raw Beacon_Type_1 Data"] + header
                csv_file_path = "saved_data/Beacon_Type_1_CSV.csv"
            elif(Beacon_Type == "2"):
                header_row = ["SN","Remarks","Raw Beacon_Type_2 Data"] + header   
                csv_file_path = "saved_data/Beacon_Type_2_CSV.csv"
            sheet.append(header_row)
                
    # Calculate next row and insert values
    next_row = sheet.max_row + 1

    # Write formula in column A and data in column B
    if next_row == 2:  # For the second row, insert the initial number '1'
        sheet.cell(row=next_row, column=1).value = 1
    else:  # For subsequent rows, use the formula
        sheet.cell(row=next_row, column=1).value = f"=A{next_row-1}+1"

    sheet.cell(row=next_row, column=2).value = profile 
    sheet.cell(row=next_row, column=3).value = raw_beacon_hex
    
    # Insert data into the same row as raw_beacon_hex
    for i, value in enumerate(data, start=4):  # Starting from column D (4th column)
        sheet.cell(row=next_row, column=i).value = value

    # Save the Excel file
    workbook.save(file_name)
    
    if(Beacon_Type == "1"):
        csv_file_path = "saved_data/Beacon_Type_1_CSV.csv"
    elif(Beacon_Type == "2"):
        csv_file_path = "saved_data/Beacon_Type_2_CSV.csv"
    
    # Opening a output csv file in write mode
    OutputCsvFile = csv.writer(open(csv_file_path, 'w'), delimiter=",")

    # Traversing in each row of the worshsheet
    for eachrow in sheet.rows:
        # Writing data of the excel file into the result csv file row-by-row
        OutputCsvFile.writerow([cell.value for cell in eachrow])
    
    beacon_data.destroy()
    show_alert()

def hk_save(profile, array_with_newline, header, data, hk_data):
    print(f"{profile}\n{array_with_newline}\n{header}\nData==={data}")
    
    # Create or load the Excel file
    file_name = "saved_data/HK_PKT_PARSER.xlsx"
    try:
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        if header:
            header_row = ["SN","Remarks","Raw HK Data"] + header
            sheet.append(header_row)
    
    # Initialize an iterator for array_with_newline
    array_iterator = iter(array_with_newline)

    for result in data:
        # Calculate next row and insert values
        next_row = sheet.max_row + 1
        # Write formula in column A and data in column B
        if next_row == 2:  # For the second row, insert the initial number '1'
            sheet.cell(row=next_row, column=1).value = 1
        else:  # For subsequent rows, use the formula
            sheet.cell(row=next_row, column=1).value = f"=A{next_row-1}+1"
        
        # Insert data into the same row as raw_address_hex
        for i, value in enumerate(result, start=4):  # Starting from column D (4th column)
            sheet.cell(row=next_row, column=2).value = profile 
            sheet.cell(row=next_row, column=i).value = value
            
        # Fetch data from array_with_newline and insert into column 3
        try:
            array_value = next(array_iterator)
            sheet.cell(row=next_row, column=3).value = array_value
        except StopIteration:
            pass  # Handle the case where array_with_newline is exhausted
        
        # Save the Excel file
        workbook.save(file_name)
    
        # Opening a output csv file in write mode
        OutputCsvFile = csv.writer(open("saved_data/HK_PKT_PARSER_CSV.csv", 'w'), delimiter=",")

        # Traversing in each row of the worshsheet
        for eachrow in sheet.rows:
            # Writing data of the excel file into the result csv file row-by-row
            OutputCsvFile.writerow([cell.value for cell in eachrow])
    
    hk_data.destroy()
    show_alert()
    
def hk_chart(headers, data_array):
    print(f"headers -- > {headers} Data -- > {data_array}")
    
    # Extracting Hour, Min, and Sec values from all rows and converting to time strings
    time_values = [f"{r[1]}:{r[2]}:{r[3]}" for r in data_array]

    # Extracting data
    solar_current = [row[4:9] for row in data_array]  # Indices 4 to 8 for 'SOLAR_P1_CS' to 'SOLAR_P5_CS'
    solar_voltage = [row[18:23] for row in data_array]  # Indices 18 to 22 for 'SP1_VOLT' to 'SP5_VOLT'
    solar_temperature = [row[26:32] for row in data_array]  # Indices 26 to 31 for '-Y_Temp' to '+X_Temp'

    batt_current = [row[10:18] for row in data_array]
    batt_voltage = [row[24:26] for row in data_array]
    obc_temperature = [row[32:35] for row in data_array]

    # Creating subplots for SOLAR Current, SOLAR Voltage, and SOLAR Temperature in a single figure with a larger size
    fig, (ax1, ax2, ax3) = plt.subplots(3, 1, figsize=(12, 10))
    
    # Plotting SOLAR Current
    ax1.set_title('SOLAR Current over Time')
    for i in range(len(solar_current[0])):
        ax1.plot(range(len(time_values)), [row[i] for row in solar_current], marker='o', linestyle='-', label=f'SOLAR_P{i + 1}_CS')
    ax1.set_ylabel('SOLAR Current')
    ax1.set_xticks(range(len(time_values)))
    ax1.set_xticklabels(time_values, rotation=45)
    ax1.legend()
    mplcursors.cursor(ax1.get_lines(), hover=True)

    # Plotting SOLAR Voltage
    ax2.set_title('SOLAR Voltage over Time')
    for i in range(len(solar_voltage[0])):
        ax2.plot(range(len(time_values)), [row[i] for row in solar_voltage], marker='x', linestyle='-', label=f'SP{i + 1}_VOLT')
    ax2.set_ylabel('SOLAR Voltage')
    ax2.set_xticks(range(len(time_values)))
    ax2.set_xticklabels(time_values, rotation=45)
    ax2.legend()
    mplcursors.cursor(ax2.get_lines(), hover=True)

    # Plotting SOLAR Temperature
    temperature_names = ['-Y_TEMP', '+Y_TEMP', '-Z_TEMP', '+Z_TEMP', '-X_TEMP', '+X_TEMP']
    ax3.set_title('SOLAR Temperature over Time')
    for i in range(len(solar_temperature[0])):
        ax3.plot(range(len(time_values)), [row[i] for row in solar_temperature], marker='^', linestyle='-', label=temperature_names[i])
    ax3.set_ylabel('SOLAR Temperature')
    ax3.set_xlabel('Time (Hour:Min:Sec)')
    ax3.set_xticks(range(len(time_values)))
    ax3.set_xticklabels(time_values, rotation=45)
    ax3.legend()
    mplcursors.cursor(ax3.get_lines(), hover=True)
    
    plt.tight_layout()
    
    # Creating subplots for batt_current, batt_voltage, and obc_temperature in a separate figure
    fig2, (ax4, ax5, ax6) = plt.subplots(3, 1, figsize=(16, 10))
    
    # Plotting batt_current
    batt_names = ['BATT_CS', 'RAW_CS', 'RESET_3V3_CS', '3V3_1_CS', '3V3_2_CS', '+5V_CS', 'UNREG1_CS', 'UNREG2_CS']
    ax4.set_title('Batt Current over Time')
    for i in range(len(batt_current[0])):
        ax4.plot(range(len(time_values)), [row[i] for row in batt_current], marker='o', linestyle='-', label=batt_names[i])
    ax4.set_ylabel('Batt Current')
    ax4.set_xticks(range(len(time_values)))
    ax4.set_xticklabels(time_values, rotation=45)
    ax4.legend()
    mplcursors.cursor(ax4.get_lines(), hover=True)
    
    # Plotting batt_voltage
    batt_voltage_names = ['RAW_VOLTS', 'BATT_VOLTS']
    ax5.set_title('Batt Voltage over Time')
    for i in range(len(batt_voltage[0])):
        ax5.plot(range(len(time_values)), [row[i] for row in batt_voltage], marker='x', linestyle='-', label=batt_voltage_names[i])
    ax5.set_ylabel('Batt Voltage')
    ax5.set_xticks(range(len(time_values)))
    ax5.set_xticklabels(time_values, rotation=45)
    ax5.legend()
    mplcursors.cursor(ax5.get_lines(), hover=True)
    
    # Plotting obc_temperature
    obc_temperature_name = ['BPB_TEMP', 'BATT_TEMP', 'OBC_INT_TEMP'] 
    ax6.set_title('OBC Temperature over Time')
    for i in range(len(obc_temperature[0])):
        ax6.plot(range(len(time_values)), [row[i] for row in obc_temperature], marker='^', linestyle='-', label=obc_temperature_name[i])
    ax6.set_ylabel('OBC Temperature')
    ax6.set_xlabel('Time (Hour:Min:Sec)')
    ax6.set_xticks(range(len(time_values)))
    ax6.set_xticklabels(time_values, rotation=45)
    ax6.legend()
    mplcursors.cursor(ax6.get_lines(), hover=True)

    plt.tight_layout()
    plt.show()
    
def image_data_save(profile, image_data):
    print(f"{profile}\n{image_data}")

    # Convert hexadecimal string to bytes
    binary_data = bytes.fromhex(image_data)

    # Create an image from binary data
    image = Image.open(io.BytesIO(binary_data))

    # Save the image in PNG format
    image.save("decoded_image.png", format="PNG")

def show_alert():
    messagebox.showinfo("Successful", "Data Stored Complete")

def on_entry_focus_in_address_data(event):
    if et_address_entry.get("1.0", "end-1c") == "_Address_data_in_hex":
        et_address_entry.delete("1.0", "end-1c")
        et_address_entry.configure(show="")
        et_address_entry.configure(fg="gray")

def on_entry_focus_out_address_data(event):
    if et_address_entry.get("1.0", "end-1c") == "":
        et_address_entry.insert("1.0", "_Address_data_in_hex")
        et_address_entry.configure(fg="gray")
        
def on_entry_focus_in_beacon_data(event):
    if et_beacon_entry.get("1.0", "end-1c") == "_Beacon_data_in_hex":
        et_beacon_entry.delete("1.0", "end-1c")
        et_beacon_entry.configure(show="")
        et_beacon_entry.configure(fg="gray")
    
def on_entry_focus_out_beacon_data(event):
    if et_beacon_entry.get("1.0", "end-1c") == "":
        et_beacon_entry.insert("1.0", "_Beacon_data_in_hex")
        et_beacon_entry.configure(fg="gray")
        
def on_entry_focus_in_hk_data(event):
    if et_hk_entry.get("1.0", "end-1c") == "_HK_data_in_hex":
        et_hk_entry.delete("1.0", "end-1c")
        et_hk_entry.configure(show="")
        et_hk_entry.configure(fg="gray")
    
def on_entry_focus_out_hk_data(event):
    if et_hk_entry.get("1.0", "end-1c") == "":
        et_hk_entry.insert("1.0", "_HK_data_in_hex")
        et_hk_entry.configure(fg="gray")
        
def on_entry_focus_in_msn1_data(event):
    if et_msn1_entry.get("1.0", "end-1c") == "_MSN1_data_in_hex":
        et_msn1_entry.delete("1.0", "end-1c")
        et_msn1_entry.configure(show="")
        et_msn1_entry.configure(fg="gray")
    
def on_entry_focus_out_msn1_data(event):
    if et_msn1_entry.get("1.0", "end-1c") == "":
        et_msn1_entry.insert("1.0", "_MSN1_data_in_hex")
        et_msn1_entry.configure(fg="gray")

def on_entry_focus_in_msn2_data(event):
    if et_msn2_entry.get("1.0", "end-1c") == "_MSN2_data_in_hex":
        et_msn2_entry.delete("1.0", "end-1c")
        et_msn2_entry.configure(text="")
        et_msn2_entry.configure(fg="gray")
    
def on_entry_focus_out_msn2_data(event):
    if et_msn2_entry.get("1.0", "end-1c") == "":
        et_msn2_entry.insert("1.0", "_MSN2_data_in_hex")
        et_msn2_entry.configure(fg="gray")
        
def on_entry_focus_in_sat_log(event):
    if et_sat_log_entry.get("1.0", "end-1c") == "_Sat_LOG_in_hex":
        et_sat_log_entry.delete("1.0", "end-1c")
        et_sat_log_entry.configure(show="")
        et_sat_log_entry.configure(fg="gray")
    
def on_entry_focus_out_sat_log(event):
    if et_sat_log_entry.get("1.0", "end-1c") == "":
        et_sat_log_entry.insert("1.0", "_Sat_LOG_in_hex")
        et_sat_log_entry.configure(fg="gray")
        
def on_entry_focus_in_name(event):
    if et_operator_name.get() == "_Full_Name_":
        et_operator_name.delete(0, END)
        et_operator_name.configure(show="")
        et_operator_name.configure(fg="gray")
        
def on_entry_focus_in_temp(event):
    if et_room_temperature.get() == "_Temperature_in_Celsius_":
        et_room_temperature.delete(0, END)
        et_room_temperature.configure(show="")
        et_room_temperature.configure(fg="gray")
        
def on_entry_focus_in_humidity(event):
    if et_room_humidity.get() == "_Humidity_in_Percentage_":
        et_room_humidity.delete(0, END)
        et_room_humidity.configure(show="")
        et_room_humidity.configure(fg="gray")

def on_entry_focus_out_name(event):
    if et_operator_name.get() == "":
        et_operator_name.insert(0, "_Full_Name_")
        et_operator_name.configure(fg="gray")
        
def on_entry_focus_out_temp(event):   
    if et_room_temperature.get() == "":
        et_room_temperature.insert(0, "_Temperature_in_Celsius_")
        et_room_temperature.configure(fg="gray")
        
def on_entry_focus_out_humidity(event):         
    if et_room_humidity.get() == "":
        et_room_humidity.insert(0, "_Humidity_in_Percentage_")
        et_room_humidity.configure(fg="gray")
        
def display_time():  
    global hour, minute, second, date, month, year
    # using the strftime() method to represent current time in string  
    hour = str(time.strftime("%H"))  
    minute = str(time.strftime("%M"))  
    second = str(time.strftime("%S"))
    date = str(time.strftime("%d"))
    month = str(time.strftime("%B"))
    year = str(time.strftime("%Y"))
    
    # configuring the text of the hour, minute, and second labels  
    hour_label.config(text = hour)  
    minute_label.config(text = minute)  
    second_label.config(text = second)
    date_label.config(text = date)  
    month_label.config(text = month)
    year_label.config(text = year)
    
    # using the after() to call the display_time() after 200 milliseconds  
    hour_label.after(200, display_time)  

def distroy_all():
    login_frame.place_forget()
    window2_operation()

def clicked():
    distroy_flag = TRUE
    if (operator_name.get() == "_Full_Name_" or room_temperature.get() == "_Temperature_in_Celsius_" or room_humidity.get() == "_Humidity_in_Percentage_"):
        distroy_flag = FALSE
    next_process(distroy_flag)

def next_process(distroy_flag):
    if (operator_name.get() == 0 or room_temperature.get() == 0 or room_humidity.get() == 0):
        distroy_flag = FALSE
    
    if distroy_flag:
        print("All the information provided")
        print(f"Operator is {operator_name.get()}")
        print(f"Room Temperature is {room_temperature.get()}")
        print(f"Room Humidity is {room_humidity.get()}")
        print(f"Hour is {hour}")
        print(f"Minute is {minute}")
        print(f"Second is {second}")
        print(f"Ready to distroy is {distroy_flag}")
        distroy_all()
    else:
        print("Please, provided all the information!")

def address_proc():
    text = et_address_entry.get("1.0", "end-1c")
    print(f"raw_address_data: {text}")
    popup.address_data_popup(text)
    
def beacon_proc():
    text = et_beacon_entry.get("1.0", "end-1c")
    print(f"raw_beacon_data: {text}")
    popup.beacon_data_popup(text)
    
def hk_proc():
    text = et_hk_entry.get("1.0", "end-1c")
    print(f"Saved text: {text}")
    popup.hk_data_popup(text)
    
def msn1_proc():
    text = et_msn1_entry.get("1.0", "end-1c")
    print(f"Saved text: {text}")
    popup.image_data_popup(text)
    
def msn2_proc():
    text = et_msn2_entry.get("1.0", "end-1c")
    print(f"Saved text: {text}")
    
def sat_log_proc():
    text = et_sat_log_entry.get("1.0", "end-1c")
    print(f"Saved text: {text}")

def window2_operation():
    sat_data_frame = Frame(root, padx=10, pady=10)
    sat_data_frame.place(x=10,y=120)

    global et_address_entry
    global et_beacon_entry
    global et_hk_entry
    global et_msn1_entry
    global et_msn2_entry
    global et_sat_log_entry
    
    ######## ADDRESS DATA ####### 
    address_data_label = LabelFrame(sat_data_frame, text="Address Data", padx=2, pady=2)
    address_data_label.grid(row=0,column=0, sticky=N+W)
    enter_address_data_label=Label(address_data_label, text="Enter Addrress data here")
    enter_address_data_label.grid(row=0,column=0)
    et_address_entry = Text(address_data_label,width=95, height= 5, fg="gray", padx=5, pady=5)
    et_address_entry.insert(END,"_Address_data_in_hex")
    et_address_entry.bind("<FocusIn>", on_entry_focus_in_address_data) 
    et_address_entry.bind("<FocusOut>", on_entry_focus_out_address_data)                             
    et_address_entry.grid(row=0,column=1)
    save_btn = Button(address_data_label, text = 'Process', command=address_proc)
    save_btn.grid(row =0, column=3)
    
    ########### Beacon Data ##########
    beacon_data_label = LabelFrame(sat_data_frame,text="Beacon Data", padx=2, pady=2)
    beacon_data_label.grid(row=1,column=0, sticky=N+W)
    enter_beacon_data_label = Label(beacon_data_label,text="Enter Beacon Data here")
    enter_beacon_data_label.grid(row=0,column=0)
    et_beacon_entry = Text(beacon_data_label,width=96, height= 5, fg="gray", padx=5, pady=5)
    et_beacon_entry.insert(END,"_Beacon_data_in_hex")
    et_beacon_entry.bind("<FocusIn>", on_entry_focus_in_beacon_data) 
    et_beacon_entry.bind("<FocusOut>", on_entry_focus_out_beacon_data)   
    et_beacon_entry.grid(row=0,column=1)
    save_btn = Button(beacon_data_label, text = 'Process', command=beacon_proc)
    save_btn.grid(row =0, column=3)
    
    ########### HK Data ##########
    hk_data_label = LabelFrame(sat_data_frame,text="HK Data", padx=2, pady=2)
    hk_data_label.grid(row=2,column=0,sticky=N+W)
    enter_hk_data_label = Label(hk_data_label,text="Enter HK Data here")
    enter_hk_data_label.grid(row=0,column=0)
    et_hk_entry = Text(hk_data_label, width = 100, height= 5, fg="gray", padx=5, pady=5)
    et_hk_entry.insert(END,"_HK_data_in_hex")
    et_hk_entry.bind("<FocusIn>", on_entry_focus_in_hk_data) 
    et_hk_entry.bind("<FocusOut>", on_entry_focus_out_hk_data)   
    et_hk_entry.grid(row=0,column=1)
    save_btn = Button(hk_data_label, text = 'Process', command=hk_proc)
    save_btn.grid(row =0, column=3)
    
    ######## MSN1 DATA ####### 
    msn1_data_label = LabelFrame(sat_data_frame, text="MSN1 Data", padx=2, pady=2)
    msn1_data_label.grid(row=3,column=0, sticky=N+W)
    enter_msn1_data_label=Label(msn1_data_label, text="Enter MSN1 data here")
    enter_msn1_data_label.grid(row=0,column=0)
    et_msn1_entry = Text(msn1_data_label, width = 98, height= 5, fg="gray", padx=5, pady=5)
    et_msn1_entry.insert(END,"_MSN1_data_in_hex")
    et_msn1_entry.bind("<FocusIn>", on_entry_focus_in_msn1_data) 
    et_msn1_entry.bind("<FocusOut>", on_entry_focus_out_msn1_data)                             
    et_msn1_entry.grid(row=0,column=1)
    save_btn = Button(msn1_data_label, text = 'Process', command=msn1_proc)
    save_btn.grid(row =0, column=3)
    
    ########### MSN2 Data ##########
    msn2_data_label = LabelFrame(sat_data_frame,text="MSN2 Data", padx=2, pady=2)
    msn2_data_label.grid(row=4,column=0, sticky=N+W)
    enter_msn2_data_label = Label(msn2_data_label,text="Enter MSN2 Data here")
    enter_msn2_data_label.grid(row=0,column=0)
    et_msn2_entry = Text(msn2_data_label, width = 98, height= 5, fg="gray", padx=5, pady=5)
    et_msn2_entry.insert(END,"_MSN2_data_in_hex")
    et_msn2_entry.bind("<FocusIn>", on_entry_focus_in_msn2_data) 
    et_msn2_entry.bind("<FocusOut>", on_entry_focus_out_msn2_data)   
    et_msn2_entry.grid(row=0,column=1)
    save_btn = Button(msn2_data_label, text = 'Process', command=msn2_proc)
    save_btn.grid(row =0, column=3)
    
    ########### Sat Log  ##########
    sat_log_label = LabelFrame(sat_data_frame,text="Sat LOG", padx=2, pady=2)
    sat_log_label.grid(row=5,column=0, sticky=N+W)
    enter_sat_log_label = Label(sat_log_label,text="Enter Sat LOG Data here")
    enter_sat_log_label.grid(row=0,column=0)
    et_sat_log_entry = Text(sat_log_label, width = 96, height= 5, fg="gray", padx=5, pady=5)
    et_sat_log_entry.insert(END,"_Sat_LOG_in_hex")
    et_sat_log_entry.bind("<FocusIn>", on_entry_focus_in_sat_log) 
    et_sat_log_entry.bind("<FocusOut>", on_entry_focus_out_sat_log)   
    et_sat_log_entry.grid(row=0,column=1)
    save_btn = Button(sat_log_label, text = 'Process', command=sat_log_proc)
    save_btn.grid(row =0, column=3)

if __name__ == "__main__":
    root = Tk()  # root as instance of TK() class
    root.title("Munal Ground Station Software")
    root.geometry("1024x768") # ("width x height)
    root.resizable(0,0)

    root_frame = Frame(root)
    root_frame.pack(anchor="nw", fill=X)

    open_image = Image.open("MUNAL.jpeg").resize((100,100))
    photo_image = ImageTk.PhotoImage(open_image)
    photo_image = ImageTk.PhotoImage(open_image)
    photo_image = ImageTk.PhotoImage(open_image)
    photo_image = ImageTk.PhotoImage(open_image)
    photo_label = Label(root_frame,image=photo_image)
    photo_label.pack(side="right", padx=10)

    Label(root_frame,text="Welcome to Munal Ground Station", fg ="red", font="Arial 28 bold", pady=5).pack()
    Label(root_frame,text="Antarikchya Pratisthan Nepal", fg ="white", font="Arial 14 italic").pack()
    Label(root_frame,text="LAZIMPAT, KATHMANDU", fg ="white", font="Arial 14 italic", pady=5).pack()

    canvas = Canvas(root, width=2500)
    canvas.pack(anchor="nw")
    canvas_width = canvas.winfo_reqwidth()
    canvas_width_v1 = canvas.winfo_reqwidth()/10
    canvas_width_v2 = canvas.winfo_reqwidth()/5
    canvas.create_line(0,5,canvas_width,5, fill="red", width=2)

    operator_name = StringVar()
    room_temperature = StringVar()
    room_humidity = StringVar()

    login_frame = Frame(root)
    login_frame.place(x=canvas_width_v1,y=150)

    operation_info_label = LabelFrame(login_frame, text="LOGIN AS OPERATOR")
    operation_info_label.grid(row=0,column=0, padx=20, pady=20)

    Label(operation_info_label, text="Operator Name", padx=10, pady=10).grid(row=1,column=0)
    et_operator_name = Entry(operation_info_label, width = 25, textvariable=operator_name,fg="gray")
    et_operator_name.insert(0,"_Full_Name_")
    et_operator_name.bind("<FocusIn>", on_entry_focus_in_name)
    et_operator_name.bind("<FocusOut>", on_entry_focus_out_name)
    et_operator_name.grid(row=1,column=1, padx=10, pady=10)

    Label(operation_info_label, text="Room Temperature", padx=10, pady=10).grid(row=2,column=0)
    et_room_temperature = Entry(operation_info_label, width = 25, textvariable=room_temperature,fg="gray")
    et_room_temperature.insert(0, "_Temperature_in_Celsius_")
    et_room_temperature.bind("<FocusIn>", on_entry_focus_in_temp)
    et_room_temperature.bind("<FocusOut>", on_entry_focus_out_temp)
    et_room_temperature.grid(row=2,column=1, padx=10, pady=10)

    Label(operation_info_label, text="Room Humidity", padx=10, pady=10).grid(row=3,column=0)
    et_room_humidity = Entry(operation_info_label, width = 25, textvariable=room_humidity,fg="gray")
    et_room_humidity.insert(0,"_Humidity_in_Percentage_")
    et_room_humidity.bind("<FocusIn>", on_entry_focus_in_humidity)
    et_room_humidity.bind("<FocusOut>", on_entry_focus_out_humidity)
    et_room_humidity.grid(row=3,column=1, padx=10, pady=10)

    login_btn = Button(operation_info_label,text="____Start____", command=clicked).grid(row=4, column=1, padx=10, pady=10)

    time_frame = Frame(operation_info_label)
    time_frame.grid(row=4, column=0)
    
  # defining some labels to display the time in the "HH:MM:SS AM/PM" format  
    hour_label = Label(time_frame, text = "00", font = ("radioland", "18"), fg = "#00D2FF")  
    colon_label_one = Label(time_frame, text = ":", font = ("radioland", "18"), fg = "#00D2FF")  
    minute_label = Label(time_frame, text = "00", font = ("radioland", "18"), fg = "#00D2FF")  
    colon_label_two = Label(time_frame, text = ":", font = ("radioland", "18"), fg = "#00D2FF")  
    second_label = Label(time_frame, text = "00", font = ("radioland", "18"), fg = "#00D2FF") 
    date_label = Label(time_frame, text = "00", font = ("radioland", "18"), fg = "#00D2FF") 
    month_label = Label(time_frame, text = "00", font = ("radioland", "18"), fg = "#00D2FF") 
    year_label = Label(time_frame, text = "00", font = ("radioland", "18"), fg = "#00D2FF") 
    
    hour_label.grid(row = 0, column = 0, padx = 5, pady = 5)
    colon_label_one.grid(row = 0, column = 1, padx = 5, pady = 5)
    minute_label.grid(row = 0, column = 2, padx = 5, pady = 5)
    colon_label_two.grid(row = 0, column = 3, padx = 5, pady = 5)
    second_label.grid(row = 0, column = 4, padx = 5, pady = 5)
    date_label.grid(row = 1, column = 0, padx = 5, pady = 5)
    month_label.grid(row = 1, column= 1, columnspan= 2, padx = 5, pady = 5)
    year_label.grid(row = 1, column= 3, columnspan= 2, padx = 5, pady = 5)
    display_time()
    
root.mainloop() # event loop or main loop